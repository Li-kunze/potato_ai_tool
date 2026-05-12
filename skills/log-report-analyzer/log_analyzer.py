#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
用户登录及功能访问分析报告生成器
支持配置文件驱动，可复用于不同据点
"""

import os
import sys
import argparse
import yaml
import psycopg2
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 默认配置
DEFAULT_CONFIG = {
    'database': {
        'host': '10.191.41.157',
        'port': '5433',
        'name': 'flexplanner_all',
        'user': 'flexplanner',
        'password': 'flexplanner@2025'
    },
    'user_mapping': {
        'ADMIN': None,
        'APS01': '排产员',
        'DDMRP001': '缓冲计划员',
        '01857': '周天益',
        '05568': '平佳威',
        '01019': '何君军',
        '05723': '王菲',
        '05192': '黄珉',
        '03047': '袁桂莲',
        '06490': '于琴',
        '05310': '陈彤',
        '01053': '潘徐芳',
        '01131': '郝要进',
        '07778': '谢梦杰',
        '06794': '王鑫月',
        '01089': '费敏敏',
        '01261': '杨旭明',
        '02181': '吴柯达',
        '02904': '励佳熹',
        '03299': '许乙杰',
        '03853': '代婷',
        '04967': '胡莎莎',
        '05626': '赵丹辉',
        '01075': '傅勇',
        '01012': '卢迪',
        '01037': '胡慧',
        '07652': '符一超',
        '01033': '丁红霞',
        '09213': '赵雨航',
        '01086': '徐骏'
    },
    'excluded_url_prefixes': [
        "/flexplanner/plugins",
        "/flexplanner/webetl",
        "/flexplanner/error",
        "/flexplanner/excel",
        "/flexplanner/ddmrp/getDdmrpSysUserDefInfo.json",
        "/flexplanner/ddmrp/getItemClassTree.json",
        "/flexplanner/items/api/getOriginalExpandingItems.json",
        "/flexplanner/ddmrp/getFutureDltDemandQtySetting.json",
        "/flexplanner/ddmrp/getBufferModel.json",
        "/flexplanner/ddmrp/getDdmrpSysUserDefInfoList.json",
        "/flexplanner/ddmrp/getWorkShopAndSupplierTree.json",
        "/flexplanner/ddmrp/getDdmrpDropDownData.json",
        "/flexplanner/ddmrp/getBufferDefineDropDownData.json",
        "/flexplanner/apsSemiWipRequirement/getDatetime.json",
        "/flexplanner/ddmrp/getResourcesTree.json",
        "/flexplanner/ddmrp/getLocationCounter.json"
    ],
    'excluded_function_prefixes': [
        "CommonController.",
        "IndexController.",
        "SchedulerLogController.",
        "TransJobInfoController.",
        "TransController.",
        "UserController.",
        "FileSystemController",
        "JobController.",
        "PageSettingsController.",
        "GlobalErrorController.",
        "SchedulerHomeController",
        "MenuManageController.",
        "MainPageController.",
        "TemplateController.",
        "I18nController.",
        "HomeController.",
        "ApiReRunController.",
        "SchedulerController.",
        "UserItemAuthorityController",
        "MstLocationInfoController.",
        "HomeDashboardController.",
        "MasterManageController."
    ],
    'report': {
        'days': 30,
        'top_n': 10
    }
}

CONFIG_FILE = 'log_analyzer_config.yaml'


def load_config():
    """加载配置文件，如果不存在则创建默认配置"""
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
        print(f"已加载配置文件: {CONFIG_FILE}")
        return config
    else:
        # 创建默认配置文件
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            yaml.dump(DEFAULT_CONFIG, f, allow_unicode=True, sort_keys=False)
        print(f"已创建默认配置文件: {CONFIG_FILE}")
        print("请根据需要修改配置文件后重新运行")
        return DEFAULT_CONFIG


def get_db_connection(config):
    """建立数据库连接"""
    db_config = config['database']
    return psycopg2.connect(
        host=db_config['host'],
        port=db_config['port'],
        database=db_config['name'],
        user=db_config['user'],
        password=db_config['password']
    )


def fetch_data(conn, sql, params):
    """执行SQL查询并返回DataFrame"""
    return pd.read_sql(sql, conn, params=params)


def filter_common_methods(df, config):
    """过滤掉共通方法"""
    excluded_prefixes = config['excluded_url_prefixes']
    excluded_function_prefixes = config['excluded_function_prefixes']
    
    # 按URL前缀过滤
    for prefix in excluded_prefixes:
        df = df[~df['访问URL'].str.startswith(prefix, na=False)]
    
    # 按功能名称前缀过滤
    for func_prefix in excluded_function_prefixes:
        df = df[~df['功能名称'].str.startswith(func_prefix, na=False)]
    
    return df.head(100)


def create_styled_workbook():
    """创建带样式的Excel工作簿"""
    wb = Workbook()
    
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(horizontal='left', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    return wb, {
        'header_font': header_font,
        'header_fill': header_fill,
        'header_alignment': header_alignment,
        'cell_alignment': cell_alignment,
        'number_alignment': number_alignment,
        'thin_border': thin_border
    }


def style_worksheet(ws, df, styles, sheet_name):
    """为工作表应用样式"""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == 1:
                cell.font = styles['header_font']
                cell.fill = styles['header_fill']
                cell.alignment = styles['header_alignment']
            else:
                if c_idx in [3, 4]:
                    cell.alignment = styles['number_alignment']
                else:
                    cell.alignment = styles['cell_alignment']
            
            cell.border = styles['thin_border']
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 25


def generate_sql(days):
    """生成SQL查询语句"""
    end_date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    start_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    
    login_detail_sql = """
    SELECT 
        operate_user AS "用户名",
        DATE(operate_date) AS "登录日期",
        COUNT(*) AS "登录次数"
    FROM temp_login_log
    WHERE operate_date >= %s 
        AND operate_date < %s::date + INTERVAL '1 day'
        AND operate_user IS NOT NULL
        AND operate_user != 'ADMIN'
    GROUP BY operate_user, DATE(operate_date)
    ORDER BY "登录日期" DESC, "登录次数" DESC, "用户名"
    """
    
    login_top10_sql = """
    SELECT 
        operate_user AS "用户名",
        COUNT(*) AS "总登录次数"
    FROM temp_login_log
    WHERE operate_date >= %s 
        AND operate_date < %s::date + INTERVAL '1 day'
        AND operate_user IS NOT NULL
        AND operate_user != 'ADMIN'
    GROUP BY operate_user
    ORDER BY "总登录次数" DESC
    LIMIT 10
    """
    
    operation_sql = """
    SELECT 
        request_uri AS "访问URL",
        COALESCE(operation, '未知功能') AS "功能名称",
        COUNT(*) AS "访问次数",
        ROUND(AVG(request_time), 2) AS "平均响应时间(ms)"
    FROM temp_operation_log
    WHERE operate_date >= %s 
        AND operate_date < %s::date + INTERVAL '1 day'
        AND request_uri IS NOT NULL
        AND site_id = 'SUPU'
        AND operate_user != 'ADMIN'
    GROUP BY request_uri, operation
    ORDER BY "访问次数" DESC
    LIMIT 200
    """
    
    top_user_sql = """
    SELECT DISTINCT ON (request_uri)
        request_uri,
        operate_user,
        COUNT(*) as user_count
    FROM temp_operation_log
    WHERE operate_date >= %s 
        AND operate_date < %s::date + INTERVAL '1 day'
        AND request_uri IS NOT NULL
        AND site_id = 'SUPU'
        AND operate_user != 'ADMIN'
    GROUP BY request_uri, operate_user
    ORDER BY request_uri, user_count DESC
    """
    
    return start_date, end_date, login_detail_sql, login_top10_sql, operation_sql, top_user_sql


def generate_report(config=None):
    """生成Excel报告"""
    if config is None:
        config = load_config()
    
    days = config['report']['days']
    user_mapping = config['user_mapping']
    
    start_date, end_date, login_detail_sql, login_top10_sql, operation_sql, top_user_sql = generate_sql(days)
    
    print(f"开始生成报告...")
    print(f"时间范围: {start_date} 至 {end_date}")
    
    conn = None
    try:
        print("正在连接数据库...")
        conn = get_db_connection(config)
        
        # 查询登录明细
        print("正在查询登录日志明细...")
        login_detail_df = fetch_data(conn, login_detail_sql, (start_date, end_date))
        print(f"  - 获取到 {len(login_detail_df)} 条登录明细记录")
        
        print("正在映射用户名...")
        login_detail_df['用户名'] = login_detail_df['用户名'].map(user_mapping)
        login_detail_df = login_detail_df[login_detail_df['用户名'].notna()]
        print(f"  - 映射后剩余 {len(login_detail_df)} 条记录（已排除ADMIN）")
        
        # 查询登录TOP10
        print("正在查询登录TOP10...")
        login_top10_df = fetch_data(conn, login_top10_sql, (start_date, end_date))
        login_top10_df['用户名'] = login_top10_df['用户名'].map(user_mapping)
        login_top10_df = login_top10_df[login_top10_df['用户名'].notna()]
        print(f"  - 获取到 {len(login_top10_df)} 条TOP10记录")
        
        # 查询操作日志
        print("正在查询操作日志...")
        operation_df = fetch_data(conn, operation_sql, (start_date, end_date))
        print(f"  - 获取到 {len(operation_df)} 条原始记录")
        
        # 过滤共通方法
        print("正在过滤共通方法...")
        operation_df = filter_common_methods(operation_df, config)
        print(f"  - 过滤后剩余 {len(operation_df)} 条业务功能记录")
        print(f"  - 排除的URL前缀: {', '.join(config['excluded_url_prefixes'][:5])}...")
        print(f"  - 排除的功能前缀: {', '.join(config['excluded_function_prefixes'][:5])}...")
        
        # 查询每个URL使用最多的用户
        print("正在查询每个功能使用最多的用户...")
        top_user_df = fetch_data(conn, top_user_sql, (start_date, end_date))
        top_user_df['operate_user'] = top_user_df['operate_user'].map(user_mapping)
        top_user_df = top_user_df[top_user_df['operate_user'].notna()]
        url_to_user = dict(zip(top_user_df['request_uri'], top_user_df['operate_user']))
        operation_df['使用最多的人名'] = operation_df['访问URL'].map(url_to_user)
        print(f"  - 已添加使用最多的人名列")
        
        # 创建Excel工作簿
        print("正在生成Excel文件...")
        wb, styles = create_styled_workbook()
        
        # Sheet 1: 用户登录明细
        ws1 = wb.active
        ws1.title = "用户登录明细"
        style_worksheet(ws1, login_detail_df, styles, "用户登录明细")
        
        # Sheet 2: 登录次数TOP10
        ws2 = wb.create_sheet(title="登录次数TOP10")
        style_worksheet(ws2, login_top10_df, styles, "登录次数TOP10")
        
        # Sheet 3: 高频功能分析
        ws3 = wb.create_sheet(title="高频功能分析")
        style_worksheet(ws3, operation_df, styles, "高频功能分析")
        
        # 生成文件名
        timestamp = datetime.now().strftime("%H%M%S")
        filename = f"用户登录及功能访问分析报告_{start_date}_至_{end_date}_{timestamp}.xlsx"
        
        # 保存文件
        wb.save(filename)
        print(f"\n[OK] 报告生成成功!")
        print(f"  文件名: {filename}")
        print(f"  保存路径: {os.path.abspath(filename)}")
        
        # 输出统计摘要
        print(f"\n=== 数据摘要 ===")
        print(f"【用户登录情况 - 近{days}天】")
        print(f"  - 统计时间: {start_date} 至 {end_date}")
        print(f"  - 登录明细记录数: {len(login_detail_df)} 条")
        print(f"  - 登录用户数: {login_detail_df['用户名'].nunique()} 人")
        print(f"  - 总登录次数: {login_detail_df['登录次数'].sum()} 次")
        
        print(f"\n【登录次数TOP10】")
        for idx, row in login_top10_df.iterrows():
            print(f"  {idx+1}. {row['用户名']}: {row['总登录次数']} 次")
        
        print(f"\n【高频功能分析 TOP{len(operation_df)}】")
        for idx, row in operation_df.head(10).iterrows():
            print(f"  {idx+1}. {row['功能名称']}")
        
        return filename
        
    except Exception as e:
        print(f"\n[ERROR] 生成报告时出错: {str(e)}")
        raise
    finally:
        if conn:
            conn.close()
            print("\n数据库连接已关闭")


def main():
    parser = argparse.ArgumentParser(description='用户登录及功能访问分析报告生成器')
    parser.add_argument('--config', help='配置文件路径', default=CONFIG_FILE)
    parser.add_argument('--db-host', help='数据库主机')
    parser.add_argument('--db-port', help='数据库端口')
    parser.add_argument('--db-name', help='数据库名称')
    parser.add_argument('--db-user', help='数据库用户')
    parser.add_argument('--db-password', help='数据库密码')
    parser.add_argument('--days', type=int, help='统计天数', default=30)
    
    args = parser.parse_args()
    
    # 加载配置
    if os.path.exists(args.config):
        with open(args.config, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
    else:
        config = DEFAULT_CONFIG.copy()
    
    # 命令行参数覆盖配置
    if args.db_host:
        config['database']['host'] = args.db_host
    if args.db_port:
        config['database']['port'] = args.db_port
    if args.db_name:
        config['database']['name'] = args.db_name
    if args.db_user:
        config['database']['user'] = args.db_user
    if args.db_password:
        config['database']['password'] = args.db_password
    if args.days:
        config['report']['days'] = args.days
    
    generate_report(config)


if __name__ == '__main__':
    main()
